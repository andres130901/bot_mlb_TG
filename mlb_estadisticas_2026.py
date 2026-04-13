import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.chart import Series

# ==================== DATOS ACTUALIZADOS AL 13 DE ABRIL 2026 ====================
data = [
    # AL East
    {"Equipo": "New York Yankees", "División": "AL East", "Pos": 1, "W": 8, "L": 7, "PCT": 0.533, "GB": 0.0, "STRK": "L5", "L10": "4-6", "Casa": "3-3", "Visitante": "5-4", "RS": 65, "RA": 44, "Over": 7, "Under": 7, "Push": 1},
    {"Equipo": "Tampa Bay Rays", "División": "AL East", "Pos": 2, "W": 8, "L": 7, "PCT": 0.533, "GB": 0.0, "STRK": "W3", "L10": "6-4", "Casa": "4-2", "Visitante": "4-5", "RS": 70, "RA": 80, "Over": 10, "Under": 3, "Push": 2},
    {"Equipo": "Baltimore Orioles", "División": "AL East", "Pos": 3, "W": 8, "L": 7, "PCT": 0.533, "GB": 0.0, "STRK": "W2", "L10": "6-4", "Casa": "5-4", "Visitante": "3-3", "RS": 60, "RA": 59, "Over": 8, "Under": 7, "Push": 0},
    {"Equipo": "Toronto Blue Jays", "División": "AL East", "Pos": 4, "W": 6, "L": 9, "PCT": 0.400, "GB": 2.0, "STRK": "L2", "L10": "2-8", "Casa": "6-6", "Visitante": "0-3", "RS": 57, "RA": 82, "Over": 6, "Under": 8, "Push": 1},
    {"Equipo": "Boston Red Sox", "División": "AL East", "Pos": 5, "W": 6, "L": 9, "PCT": 0.400, "GB": 2.0, "STRK": "L1", "L10": "4-6", "Casa": "3-4", "Visitante": "3-5", "RS": 55, "RA": 68, "Over": 7, "Under": 8, "Push": 0},

    # AL Central
    {"Equipo": "Cleveland Guardians", "División": "AL Central", "Pos": 1, "W": 9, "L": 7, "PCT": 0.563, "GB": 0.0, "STRK": "W2", "L10": "7-3", "Casa": "5-2", "Visitante": "4-5", "RS": 79, "RA": 71, "Over": 8, "Under": 7, "Push": 1},
    {"Equipo": "Minnesota Twins", "División": "AL Central", "Pos": 2, "W": 9, "L": 7, "PCT": 0.563, "GB": 0.0, "STRK": "W1", "L10": "6-4", "Casa": "4-2", "Visitante": "5-5", "RS": 58, "RA": 67, "Over": 7, "Under": 9, "Push": 0},
    {"Equipo": "Detroit Tigers", "División": "AL Central", "Pos": 3, "W": 7, "L": 9, "PCT": 0.438, "GB": 2.0, "STRK": "L3", "L10": "4-6", "Casa": "3-4", "Visitante": "4-5", "RS": 62, "RA": 70, "Over": 8, "Under": 7, "Push": 1},
    {"Equipo": "Kansas City Royals", "División": "AL Central", "Pos": 4, "W": 6, "L": 9, "PCT": 0.400, "GB": 3.0, "STRK": "L1", "L10": "5-5", "Casa": "4-3", "Visitante": "2-6", "RS": 55, "RA": 64, "Over": 6, "Under": 8, "Push": 1},
    {"Equipo": "Chicago White Sox", "División": "AL Central", "Pos": 5, "W": 5, "L": 10, "PCT": 0.333, "GB": 4.0, "STRK": "W2", "L10": "3-7", "Casa": "2-5", "Visitante": "3-5", "RS": 48, "RA": 78, "Over": 5, "Under": 9, "Push": 1},

    # AL West
    {"Equipo": "Los Angeles Angels", "División": "AL West", "Pos": 1, "W": 9, "L": 6, "PCT": 0.600, "GB": 0.0, "STRK": "W4", "L10": "7-3", "Casa": "5-2", "Visitante": "4-4", "RS": 68, "RA": 55, "Over": 9, "Under": 5, "Push": 1},
    {"Equipo": "Houston Astros", "División": "AL West", "Pos": 2, "W": 8, "L": 7, "PCT": 0.533, "GB": 1.0, "STRK": "L2", "L10": "6-4", "Casa": "5-3", "Visitante": "3-4", "RS": 72, "RA": 66, "Over": 8, "Under": 7, "Push": 0},
    {"Equipo": "Seattle Mariners", "División": "AL West", "Pos": 3, "W": 7, "L": 8, "PCT": 0.467, "GB": 2.0, "STRK": "W1", "L10": "5-5", "Casa": "4-3", "Visitante": "3-5", "RS": 59, "RA": 61, "Over": 6, "Under": 8, "Push": 1},
    {"Equipo": "Texas Rangers", "División": "AL West", "Pos": 4, "W": 7, "L": 8, "PCT": 0.467, "GB": 2.0, "STRK": "L1", "L10": "4-6", "Casa": "3-4", "Visitante": "4-4", "RS": 65, "RA": 72, "Over": 7, "Under": 7, "Push": 1},
    {"Equipo": "Oakland Athletics", "División": "AL West", "Pos": 5, "W": 5, "L": 10, "PCT": 0.333, "GB": 4.0, "STRK": "L4", "L10": "3-7", "Casa": "2-6", "Visitante": "3-4", "RS": 52, "RA": 81, "Over": 5, "Under": 9, "Push": 1},

    # NL East
    {"Equipo": "Atlanta Braves", "División": "NL East", "Pos": 1, "W": 10, "L": 6, "PCT": 0.625, "GB": 0.0, "STRK": "W1", "L10": "7-3", "Casa": "6-2", "Visitante": "4-4", "RS": 78, "RA": 62, "Over": 9, "Under": 6, "Push": 1},
    {"Equipo": "New York Mets", "División": "NL East", "Pos": 2, "W": 9, "L": 7, "PCT": 0.563, "GB": 1.0, "STRK": "W3", "L10": "6-4", "Casa": "5-3", "Visitante": "4-4", "RS": 71, "RA": 58, "Over": 8, "Under": 7, "Push": 1},
    {"Equipo": "Philadelphia Phillies", "División": "NL East", "Pos": 3, "W": 8, "L": 7, "PCT": 0.533, "GB": 1.5, "STRK": "L1", "L10": "5-5", "Casa": "4-3", "Visitante": "4-4", "RS": 66, "RA": 63, "Over": 7, "Under": 7, "Push": 1},
    {"Equipo": "Miami Marlins", "División": "NL East", "Pos": 4, "W": 6, "L": 9, "PCT": 0.400, "GB": 3.5, "STRK": "L3", "L10": "4-6", "Casa": "3-5", "Visitante": "3-4", "RS": 54, "RA": 73, "Over": 6, "Under": 8, "Push": 1},
    {"Equipo": "Washington Nationals", "División": "NL East", "Pos": 5, "W": 5, "L": 10, "PCT": 0.333, "GB": 4.5, "STRK": "W2", "L10": "3-7", "Casa": "2-6", "Visitante": "3-4", "RS": 49, "RA": 79, "Over": 5, "Under": 9, "Push": 1},

    # NL Central
    {"Equipo": "Pittsburgh Pirates", "División": "NL Central", "Pos": 1, "W": 9, "L": 6, "PCT": 0.600, "GB": 0.0, "STRK": "L1", "L10": "6-4", "Casa": "5-2", "Visitante": "4-4", "RS": 68, "RA": 55, "Over": 8, "Under": 6, "Push": 1},
    {"Equipo": "Chicago Cubs", "División": "NL Central", "Pos": 2, "W": 8, "L": 7, "PCT": 0.533, "GB": 1.0, "STRK": "W2", "L10": "5-5", "Casa": "4-3", "Visitante": "4-4", "RS": 72, "RA": 67, "Over": 9, "Under": 5, "Push": 1},
    {"Equipo": "Milwaukee Brewers", "División": "NL Central", "Pos": 3, "W": 8, "L": 8, "PCT": 0.500, "GB": 1.5, "STRK": "L2", "L10": "5-5", "Casa": "5-3", "Visitante": "3-5", "RS": 65, "RA": 64, "Over": 7, "Under": 8, "Push": 1},
    {"Equipo": "Cincinnati Reds", "División": "NL Central", "Pos": 4, "W": 7, "L": 8, "PCT": 0.467, "GB": 2.0, "STRK": "W1", "L10": "6-4", "Casa": "4-4", "Visitante": "3-4", "RS": 61, "RA": 70, "Over": 8, "Under": 7, "Push": 0},
    {"Equipo": "St. Louis Cardinals", "División": "NL Central", "Pos": 5, "W": 6, "L": 9, "PCT": 0.400, "GB": 3.0, "STRK": "L4", "L10": "3-7", "Casa": "3-5", "Visitante": "3-4", "RS": 58, "RA": 74, "Over": 6, "Under": 8, "Push": 1},

    # NL West
    {"Equipo": "Los Angeles Dodgers", "División": "NL West", "Pos": 1, "W": 11, "L": 4, "PCT": 0.733, "GB": 0.0, "STRK": "L1", "L10": "8-2", "Casa": "6-2", "Visitante": "5-2", "RS": 89, "RA": 51, "Over": 7, "Under": 6, "Push": 2},
    {"Equipo": "San Diego Padres", "División": "NL West", "Pos": 2, "W": 10, "L": 6, "PCT": 0.625, "GB": 1.5, "STRK": "W5", "L10": "7-3", "Casa": "5-2", "Visitante": "5-4", "RS": 74, "RA": 58, "Over": 8, "Under": 7, "Push": 1},
    {"Equipo": "Arizona Diamondbacks", "División": "NL West", "Pos": 3, "W": 8, "L": 7, "PCT": 0.533, "GB": 3.0, "STRK": "W3", "L10": "6-4", "Casa": "4-3", "Visitante": "4-4", "RS": 70, "RA": 63, "Over": 9, "Under": 5, "Push": 1},
    {"Equipo": "San Francisco Giants", "División": "NL West", "Pos": 4, "W": 7, "L": 8, "PCT": 0.467, "GB": 4.0, "STRK": "L2", "L10": "5-5", "Casa": "4-4", "Visitante": "3-4", "RS": 62, "RA": 68, "Over": 7, "Under": 7, "Push": 1},
    {"Equipo": "Colorado Rockies", "División": "NL West", "Pos": 5, "W": 5, "L": 10, "PCT": 0.333, "GB": 6.0, "STRK": "L5", "L10": "3-7", "Casa": "3-5", "Visitante": "2-5", "RS": 54, "RA": 85, "Over": 6, "Under": 8, "Push": 1},
]

df = pd.DataFrame(data)

# Cálculos adicionales
df["% Carreras Anotadas"] = (df["RS"] / (df["RS"] + df["RA"]) * 100).round(1)
df["% Carreras Permitidas"] = (df["RA"] / (df["RS"] + df["RA"]) * 100).round(1)

columnas = ["Equipo", "División", "Pos", "W", "L", "PCT", "GB", "STRK", "L10", "Casa", "Visitante",
            "RS", "RA", "% Carreras Anotadas", "% Carreras Permitidas", "Over", "Under", "Push"]
df = df[columnas]

# ==================== CREAR EXCEL CON FORMATO AVANZADO ====================
with pd.ExcelWriter("MLB_Estadisticas_2026.xlsx", engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Estadísticas MLB", index=False)
    
    # Hoja 2: Últimos 10 días (resultados de juegos)
    resultados_recientes = pd.DataFrame({
        "Fecha": ["13/04", "12/04", "11/04", "10/04", "09/04", "08/04", "07/04", "06/04", "05/04", "04/04"],
        "Partido": ["HOU 4-2 SEA", "LAD 7-3 SF", "NYY 5-4 BAL", "ATL 6-1 MIA", "CLE 8-3 CHW", "SD 5-2 ARI", "MIN 4-2 TOR", "BOS 3-2 TB", "TEX 6-5 OAK", "PIT 7-4 CIN"],
        "Ganador": ["Houston Astros", "Los Angeles Dodgers", "New York Yankees", "Atlanta Braves", "Cleveland Guardians", "San Diego Padres", "Minnesota Twins", "Boston Red Sox", "Texas Rangers", "Pittsburgh Pirates"],
        "Carreras Totales": [6, 10, 9, 7, 11, 7, 6, 5, 11, 11],
        "Over/Under": ["Under", "Over", "Over", "Under", "Over", "Under", "Under", "Under", "Over", "Over"]
    })
    resultados_recientes.to_excel(writer, sheet_name="Últimos 10 días", index=False)

    # Cargar el libro para aplicar formatos
    wb = writer.book
    ws1 = wb["Estadísticas MLB"]
    ws2 = wb["Últimos 10 días"]

    # Ancho de columnas
    for ws in [ws1, ws2]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    # Títulos en negrita y centrados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # FILTROS AUTOMÁTICOS
    ws1.auto_filter.ref = ws1.dimensions
    ws2.auto_filter.ref = ws2.dimensions

    # ==================== FORMATOS CONDICIONALES ====================
    # Racha ganadora (STRK) → Verde
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    rule_green = Rule(type="containsText", operator="containsText", text="W", dxf= DifferentialStyle(fill=green_fill))
    rule_red = Rule(type="containsText", operator="containsText", text="L", dxf= DifferentialStyle(fill=red_fill))
    ws1.conditional_formatting.add("H2:H31", rule_green)   # Columna STRK
    ws1.conditional_formatting.add("H2:H31", rule_red)

    # Over > Under → Verde en columna Over
    rule_over = Rule(type="expression", formula=["$P2>$Q2"], dxf=DifferentialStyle(fill=green_fill))
    ws1.conditional_formatting.add("P2:P31", rule_over)

    # % Carreras Anotadas > 50% → Verde claro
    light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rule_pct = Rule(type="expression", formula=["$N2>50"], dxf=DifferentialStyle(fill=light_green))
    ws1.conditional_formatting.add("N2:N31", rule_pct)

    # ==================== GRÁFICOS ====================
    # Gráfico de barras RS vs RA (Top 10 equipos)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Carreras Anotadas vs Permitidas (Top 10)"
    chart.y_axis.title = "Carreras"
    chart.x_axis.title = "Equipos"

    data_ref = Reference(ws1, min_col=12, min_row=1, max_col=13, max_row=11)  # RS y RA
    cats = Reference(ws1, min_col=1, min_row=2, max_row=11)                  # Nombres de equipos
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws1.add_chart(chart, "T2")  # Coloca el gráfico en la celda T2

print("✅ ¡Archivo creado correctamente: MLB_Estadisticas_2026.xlsx")
print("   Abrelo con Excel y tendrás todo: filtros, colores y gráficos listos.")